from PyPDFForm import FormWrapper # type: ignore # pypdfform
#import time
import pandas # type: ignore # pandas
import os
import warnings
from openpyxl import load_workbook # type: ignore # openpyxl
import xlsxwriter

#testing imports
import sample_sorter
import aux_func
import shimadzu_init
import searcher

warnings.simplefilter(action='ignore', category=FutureWarning)

def ISAR_fill(controls, batch, path):
    low1 = controls[0]
    high1 = controls[1]
    low2 = controls[-1]
    high2 = controls[-2]

    #sort through ISAR_controls.results_ISTD and find values to pass
    def get_indexes(single_control):
        for tuples in single_control.results_ISTD:
            return tuples.index('Area'), len(single_control.results_ISTD)
    
    area_index, length = get_indexes(low1)

    field_low1 = [f'A{i}_L1_1' for i in range(1,length)]
    field_high1 = [f'A{i}_L2_1' for i in range(1,length)]
    field_low2 = [f'A{i}_L1_2' for i in range(1,length)]
    field_high2 = [f'A{i}_L2_2' for i in range(1,length)]

    value_low1 = []
    value_high1 = []
    value_low2 = []
    value_high2 = []

    for tuples in low1.results_ISTD[1:]:
        value_low1.append(tuples[area_index])
    for tuples in high1.results_ISTD[1:]:
        value_high1.append(tuples[area_index])
    for tuples in low2.results_ISTD[1:]:
        value_low2.append(tuples[area_index])
    for tuples in high2.results_ISTD[1:]:
        value_high2.append(tuples[area_index])

    # #map field name to value in dictionary
    low1_dict = dict(zip(field_low1, value_low1))
    high1_dict = dict(zip(field_high1, value_high1))
    low2_dict = dict(zip(field_low2, value_low2))
    high2_dict = dict(zip(field_high2, value_high2))

    pdf = FormWrapper(path)

    pdf.fill({'Batch': batch}, adobe_mode = True)
    pdf.fill(low1_dict, adobe_mode = True)
    pdf.fill(high1_dict, adobe_mode = True)
    pdf.fill(low2_dict, adobe_mode = True)
    pdf.fill(high2_dict, adobe_mode = True)
    
    filled_pdf_path = (path)
    with open(filled_pdf_path, "wb") as output:
        output.write(pdf.read())

    print("ISAR filled and saved successfully!")
    # #calculate own -50 - 200% range unless can read from pdf
    # #don't forget about failed cases/analytes field
    #send ISAR results to txt or csv file

def output_LJ(controls, serum_controls, batch, path):
    analyte_dataframes = {}

    single_control = controls[0]

    def get_indexes(single_control):
        for tuples in single_control.results_analyte:
            return tuples.index('Conc.'), tuples.index('Name')

    conc_index, name_index = get_indexes(single_control)   
    
    for i in range(0, len(controls), 2):
        low_control = controls[i]
        high_control = controls[i+1]

        for low_result, high_result in zip(low_control.results_analyte[1:], high_control.results_analyte[1:]):
            analyte_name = low_result[name_index]
            low_conc_value = float(low_result[conc_index])
            high_conc_value = float(high_result[conc_index])

            if analyte_name not in analyte_dataframes:
                analyte_dataframes[analyte_name] = pandas.DataFrame(columns=[
                    "Batch", "CTL Low Conc.", "CTL High Conc.", "Matrix"])

            analyte_dataframes[analyte_name] = pandas.concat(
                [analyte_dataframes[analyte_name], pandas.DataFrame({"Batch": batch,
                                                                    "CTL Low Conc.": [low_conc_value], 
                                                                    "CTL High Conc.": [high_conc_value],
                                                                    "Matrix": "Blood"
                                                                    })], ignore_index=True
            )

    for i in range(0, len(serum_controls), 2):
        low_serum_control = serum_controls[i]
        high_serum_control = serum_controls[i+1]

        for low_result_serum, high_result_serum in zip(low_serum_control.results_analyte[1:], high_serum_control.results_analyte[1:]):
            analyte_name = low_result_serum[name_index]
            low_conc_value = float(low_result_serum[conc_index])
            high_conc_value = float(high_result_serum[conc_index])

            if analyte_name not in analyte_dataframes:
                analyte_dataframes[analyte_name] = pandas.DataFrame(columns=[
                    "Batch", "CTL Low Conc.", "CTL High Conc.", "Matrix"])

            analyte_dataframes[analyte_name] = pandas.concat(
                [analyte_dataframes[analyte_name], pandas.DataFrame({"Batch": batch,
                                                                    "CTL Low Conc.": [low_conc_value], 
                                                                    "CTL High Conc.": [high_conc_value],
                                                                    "Matrix": "Serum"
                                                                    })], ignore_index=True
            )

    output_path = os.path.join(path, "LJ.xlsx")

    with pandas.ExcelWriter(output_path, engine='openpyxl') as writer:
        for analyte, df in analyte_dataframes.items():
            df.to_excel(writer, sheet_name=analyte, index=False)


        # formatted_date = time.strftime("%m%d%y", time.localtime())
        # output_path = os.path.join(output_dir, f"{name}_{batch}_{formatted_date}.pdf")
        # doc1.save(output_path)
        # print(f"completed binding Batch Pack - {name}_{batch}_{formatted_date}")

def interpret_MSA(case_list):
    #find out how many analytes there are and send to fill_MSA appropriately
    #return how many copies of excel to create and the name for each one
    base_pdf = case_list[0]
    print(len(base_pdf.results_analyte))

    positive_analytes = []
    if len(base_pdf.results_analyte) > 1:
        for tuples in base_pdf.results_analyte[1:]:
            positive_analytes.append(tuples[1])

    print(positive_analytes)

    return positive_analytes

def fill_MSA(case_list, batch, MSA_path, analyte):

    base_pdf = case_list[0]
    def get_indexes(base_pdf):
        for tuples in base_pdf.results_analyte:
            return tuples.index('Area'), tuples.index('Name')
    area_index, name_index = get_indexes(base_pdf)  
    
    ISTD_peak_area = []
    analyte_peak_area = []

    for pdf in case_list:
        print(analyte)
        for tuples in pdf.results_ISTD[1:]:
            if tuples[name_index].startswith(analyte):
                ISTD_peak_area.append(tuples[area_index])
                print('found area ISTD')
        for tuples in pdf.results_analyte[1:]:
            if analyte in tuples:
                analyte_peak_area.append(tuples[area_index])

    print(ISTD_peak_area)
    print(analyte_peak_area)
    print(MSA_path)

    d = {"ISTD Peak Area": ISTD_peak_area, "Analyte Peak Area": analyte_peak_area}
    df = pandas.DataFrame(d)

    try:
        # with pandas.ExcelWriter(MSA_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        #     df.to_excel(writer, sheet_name='LF-10 MSA Worksheet', startrow=13, startcol=3, index=False)
        workbook = load_workbook(MSA_path)
        sheet = workbook['LF-10 MSA Worksheet']

        # Define the starting row and column
        startrow = 14
        startcol = 3

        # Write DataFrame to the specified location
        for i, row in df.iterrows():
            for j, value in enumerate(row):
                cell = sheet.cell(row=startrow + i + 1, column=startcol + j + 1)
                cell.value = value

        # Save the workbook
        workbook.template = True
        workbook.save(MSA_path)

    except Exception as e:
        print(e)

#import pandas as pd
#use to fill dataframe
# # Load your existing Excel file
# df = pd.read_excel('example.xlsx')

# # Make changes to specific cells
# df.at[2, 'Column_Name'] = 'New_Value'  # Modify cell in row 3, column 'Column_Name'

# # Save the modified DataFrame back to Excel
# df.to_excel('example_modified.xlsx', index=False)



#use to fill single cells 
# from openpyxl import load_workbook

# # Load your existing Excel workbook
# wb = load_workbook('example.xlsx')

# # Select the active worksheet
# ws = wb.active

# # Modify a specific cell
# ws['B3'] = 'New_Value'  # Modify cell B3

# # Save the modified workbook
# wb.save('example_modified.xlsx')


if __name__ == '__main__':
    controls_ISTD = [
[
        ('ID#', 'Name', 'Ret. Time (min)', 'Area', 'Quant Ion (m/z)', '', '', 'Mode'),
        ('1', 'Morphine-D3', '0.412', '1600408', '289.20>165.20', '', '', ''),
        ('3', 'Codeine-D3', '0.945', '1295371', '303.20>165.20', '', '', ''),
        ('5', '6-Acetylmorphine-D3', '1.102', '853868', '331.20>165.20', '', '', ''),
        ('7', 'Benzoylecgonine-D3', '1.555', '1781774', '293.20>171.20', '', '', ''),
        ('9', 'Cocaine-D3', '1.892', '2836682', '307.20>185.20', '', '', ''),
        ('11', 'Cocaethylene-D3', '2.026', '2480191', '321.20>199.20', '', '', ''),
        ('13', 'Fentanyl-D5', '2.122', '2008895', '342.20>188.20', '', '', ''),
        ('15', 'Alprazolam-D5', '2.319', '2090488', '314.20>286.20', '', '', '')
    ],[
        ('ID#', 'Name', 'Ret. Time (min)', 'Area', 'Quant Ion (m/z)', '', '', 'Mode'),
        ('1', 'Morphine-D3', '0.412', '1600408', '289.20>165.20', '', '', ''),
        ('3', 'Codeine-D3', '0.945', '1295371', '303.20>165.20', '', '', ''),
        ('5', '6-Acetylmorphine-D3', '1.102', '853868', '331.20>165.20', '', '', ''),
        ('7', 'Benzoylecgonine-D3', '1.555', '1781774', '293.20>171.20', '', '', ''),
        ('9', 'Cocaine-D3', '1.892', '2836682', '307.20>185.20', '', '', ''),
        ('11', 'Cocaethylene-D3', '2.026', '2480191', '321.20>199.20', '', '', ''),
        ('13', 'Fentanyl-D5', '2.122', '2008895', '342.20>188.20', '', '', ''),
        ('15', 'Alprazolam-D5', '2.319', '2090488', '314.20>286.20', '', '', '')
    ],[
        ('ID#', 'Name', 'Ret. Time (min)', 'Area', 'Quant Ion (m/z)', '', '', 'Mode'),
        ('1', 'Morphine-D3', '0.412', '1600408', '289.20>165.20', '', '', ''),
        ('3', 'Codeine-D3', '0.945', '1295371', '303.20>165.20', '', '', ''),
        ('5', '6-Acetylmorphine-D3', '1.102', '853868', '331.20>165.20', '', '', ''),
        ('7', 'Benzoylecgonine-D3', '1.555', '1781774', '293.20>171.20', '', '', ''),
        ('9', 'Cocaine-D3', '1.892', '2836682', '307.20>185.20', '', '', ''),
        ('11', 'Cocaethylene-D3', '2.026', '2480191', '321.20>199.20', '', '', ''),
        ('13', 'Fentanyl-D5', '2.122', '2008895', '342.20>188.20', '', '', ''),
        ('15', 'Alprazolam-D5', '2.319', '2090488', '314.20>286.20', '', '', '')
    ],[
        ('ID#', 'Name', 'Ret. Time (min)', 'Area', 'Quant Ion (m/z)', '', '', 'Mode'),
        ('1', 'Morphine-D3', '0.412', '1600408', '289.20>165.20', '', '', ''),
        ('3', 'Codeine-D3', '0.945', '1295371', '303.20>165.20', '', '', ''),
        ('5', '6-Acetylmorphine-D3', '1.102', '853868', '331.20>165.20', '', '', ''),
        ('7', 'Benzoylecgonine-D3', '1.555', '1781774', '293.20>171.20', '', '', ''),
        ('9', 'Cocaine-D3', '1.892', '2836682', '307.20>185.20', '', '', ''),
        ('11', 'Cocaethylene-D3', '2.026', '2480191', '321.20>199.20', '', '', ''),
        ('13', 'Fentanyl-D5', '2.122', '2008895', '342.20>188.20', '', '', ''),
        ('15', 'Alprazolam-D5', '2.319', '2090488', '314.20>286.20', '', '', '')
    ],
    ]
    controls = [
    sample_sorter.Sample('LOW CTL 1_0', r"G:/", 'LOW CTL 1', None, None, 
[    
        ('ID#', 'Name', 'Ret. Time (min)', 'Area', 'Quant Ion (m/z)', 'Conc.', 'Unit', 'Mode'),
        ('2', 'Morphine', '0.413', '314808', '286.20>165.20', '0.017', 'mg/L', ''),
        ('4', 'Codeine', '0.956', '224105', '300.20>165.20', '0.019', 'mg/L', ''),
        ('6', '6-Acetylmorphine', '1.107', '152560', '328.20>165.20', '1.875', 'ng/mL', ''),
        ('8', 'Benzoylecgonine', '1.554', '298989', '290.20>168.20', '0.079', 'mg/L', ''),
        ('10', 'Cocaine', '1.888', '482118', '304.20>182.20', '0.020', 'mg/L', ''),
        ('12', 'Cocaethylene', '2.020', '509149', '318.20>196.20', '0.021', 'mg/L', ''),
        ('14', 'Fentanyl', '2.118', '957656', '337.20>188.20', '5.533', 'ng/mL', ''),
        ('16', 'Alprazolam', '2.316', '328786', '309.20>281.20', '0.020', 'mg/L', '')
    ]),
    sample_sorter.Sample('HIGH CTL 1_0', r"G:/", 'HIGH CTL 1', None, None, 
[
        ('ID#', 'Name', 'Ret. Time (min)', 'Area', 'Quant Ion (m/z)', 'Conc.', 'Unit', 'Mode'),
        ('2', 'Morphine', '0.415', '161400', '286.20>165.20', '0.008', 'mg/L', ''),
        ('4', 'Codeine', '0.955', '135171', '300.20>165.20', '0.009', 'mg/L', ''),
        ('6', '6-Acetylmorphine', '1.108', '86409', '328.20>165.20', '0.910', 'ng/mL', ''),
        ('8', 'Benzoylecgonine', '1.559', '147846', '290.20>168.20', '0.039', 'mg/L', ''),
        ('10', 'Cocaine', '1.894', '244909', '304.20>182.20', '0.010', 'mg/L', ''),
        ('12', 'Cocaethylene', '2.027', '233458', '318.20>196.20', '0.010', 'mg/L', ''),
        ('14', 'Fentanyl', '2.125', '794659', '337.20>188.20', '4.852', 'ng/mL', ''),
        ('16', 'Alprazolam', '2.323', '166212', '309.20>281.20', '0.011', 'mg/L', '')
    ])
]
    batch = 111
    method = 'QTABUSE'
    path = r"C:\Users\e314883\Desktop\locked_git_repo\12786"
    #output_LJ(controls, [], batch, path)
    #ISAR_fill(controls_ISTD, batch, method)